"""
reporting/excel_builder.py — builds the final Excel report matching the
Google Sheets template: image formula, product link, name, size,
availability, discounted price, original price, image URL, and a
savings-amount formula.

No images are downloaded — only the image URL is stored, and Google
Sheets' =IMAGE() formula renders the photo from that URL after import.
"""

import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    "Фото продукта",     # A — =IMAGE(H{row})
    "Ссылка на продукт", # B
    "Имя",                # C
    "Размер",             # D
    "Наличие",            # E
    "Цена",               # F — discounted (current) price
    "Цена без скидки",    # G — compare_at_price (original price)
    "image_url",          # H — raw image URL, feeds the IMAGE() formula
    "Цена с учетом скидки (экономия)",  # I — savings amount formula
]


def rows_to_dataframe(rows: list) -> pd.DataFrame:
    """
    rows: list of flat dicts, one per (product, size) combination:
        {
            "product_url": ...,
            "name": ...,
            "size": ...,
            "available": ...,
            "price": ...,              # discounted / current price
            "compare_at_price": ...,   # original price before discount
            "image_url": ...,
        }
    """
    return pd.DataFrame(rows)


def save_csv(df: pd.DataFrame, path: str) -> None:
    df.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"Saved {len(df)} rows to {path}")


def save_excel_with_formulas(df: pd.DataFrame, path: str) -> None:
    """
    Writes the report as an .xlsx file with the exact column layout and
    formulas expected by the Google Sheets template:

        A: =IMAGE(H{row})
        B: product_url
        C: name
        D: size
        E: available (TRUE/FALSE)
        F: price (discounted)
        G: compare_at_price (original)
        H: image_url
        I: =VALUE(SUBSTITUTE(G{row};".";","))-VALUE(SUBSTITUTE(F{row};".";","))
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(COLUMNS)

    for i, row in df.iterrows():
        excel_row = i + 2  # header is row 1

        product_url = row.get("product_url", "")
        name = row.get("name", "")
        size = row.get("size", "")
        available = row.get("available", "")
        price = row.get("price", "")
        compare_at_price = row.get("compare_at_price", "")
        image_url = row.get("image_url", "")

        ws.append([
            f"=IMAGE(H{excel_row})",
            product_url,
            name,
            size,
            available,
            price,
            compare_at_price,
            image_url,
            (
                f'=VALUE(SUBSTITUTE(G{excel_row};".";","))'
                f'-VALUE(SUBSTITUTE(F{excel_row};".";","))'
            ),
        ])

    # Reasonable default column widths
    widths = [14, 45, 40, 10, 10, 10, 16, 60, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Give image cells some row height so =IMAGE() has room to render
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    wb.save(path)
    logger.info(f"Excel report saved to {path}")
